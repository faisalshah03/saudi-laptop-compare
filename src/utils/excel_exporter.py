"""Excel file generation for price comparison"""
import os
from typing import List, Dict, Any
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    raise ImportError("openpyxl not installed. Run: python3 -m pip install openpyxl")


class ExcelExporter:
    """Export price comparison data to Excel"""

    # Column definitions
    COLUMNS = [
        ('A', 'master_sku', 'Master SKU'),
        ('B', 'title', 'Title'),
        ('C', 'category', 'Category'),
        ('D', 'subtype', 'Subtype'),
        ('E', 'brand', 'Brand'),
        ('F', 'model_name', 'Model Name'),
        ('G', 'model_number', 'Model Number'),
        ('H', 'manufacturer_number', 'Manufacturer Number'),
        ('I', 'processor', 'Processor'),
        ('J', 'processor_full', 'Processor (Full Name)'),
        ('K', 'cpu_power', 'CPU Clock Speed'),
        ('L', 'ram', 'RAM'),
        ('M', 'storage', 'Storage'),
        ('N', 'graphics_card', 'Graphics Card'),
        ('O', 'ai_classification', 'AI Classification'),
        ('P', 'npu_tops', 'NPU TOPS'),
        ('Q', 'amazon_sa_price', 'Amazon.sa (SAR)'),
        ('R', 'jarir_price', 'Jarir (SAR)'),
        ('S', 'extra_price', 'Extra.com (SAR)'),
        ('T', 'noon_price', 'Noon.com (SAR)'),
        ('U', 'best_price', 'Best Price (SAR)'),
        ('V', 'best_price_platform', 'Best on Platform'),
        ('W', 'amazon_sa_link', 'Amazon Link'),
        ('X', 'jarir_link', 'Jarir Link'),
        ('Y', 'extra_link', 'Extra Link'),
        ('Z', 'noon_link', 'Noon Link'),
        ('AA', 'last_updated', 'Last Updated'),
    ]

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet

    def _get_column_index(self, field_name: str) -> int:
        """Get column number by field name."""
        for col_letter, field, _ in self.COLUMNS:
            if field == field_name:
                return column_index_from_string(col_letter)
        return None

    def _create_header_row(self, ws):
        """Create and format header row."""
        # Write headers
        for col_letter, field, header in self.COLUMNS:
            cell = ws[f'{col_letter}1']
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF', size=12)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Auto-filter
        ws.auto_filter.ref = f'A1:AA{ws.max_row}'

        # Set column widths
        column_widths = {
            'A': 18, 'B': 40, 'C': 10, 'D': 16, 'E': 12, 'F': 18, 'G': 12,
            'H': 18, 'I': 16, 'J': 26, 'K': 14, 'L': 10, 'M': 12, 'N': 18,
            'O': 16, 'P': 12, 'Q': 14, 'R': 14, 'S': 14, 'T': 14, 'U': 14,
            'V': 16, 'W': 30, 'X': 30, 'Y': 30, 'Z': 30, 'AA': 18
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Set row height
        ws.row_dimensions[1].height = 30

    def _format_data_rows(self, ws, products: List[Dict]):
        """Format and fill data rows with proper styling."""
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row_idx, product in enumerate(products, start=2):
            for col_letter, field, _ in self.COLUMNS:
                cell = ws[f'{col_letter}{row_idx}']
                value = product.get(field)

                # Format value
                if value is None or value == '':
                    cell.value = 'N/A'
                elif isinstance(value, float) and field.endswith('_price'):
                    cell.value = value
                    cell.number_format = '#,##0.00'
                elif isinstance(value, (int, float)) and not field.endswith('_price'):
                    cell.value = value
                else:
                    cell.value = str(value)

                # Apply styling
                cell.alignment = Alignment(horizontal='center' if field.endswith('_price') else 'left',
                                          vertical='center', wrap_text=True)
                cell.border = border

                # Highlight best price in green
                if field == 'best_price' and isinstance(value, (int, float)):
                    cell.fill = green_fill
                    cell.font = Font(bold=True)

                # Highlight unavailable items in red
                if value == 'N/A' or value == 'Not Listed':
                    cell.fill = red_fill

                # Alternate row coloring for readability
                if row_idx % 2 == 0 and field not in ['best_price']:
                    if cell.fill.start_color.index == '00000000':  # No fill yet
                        cell.fill = light_gray_fill

    def create_comparison_sheet(self, products: List[Dict]):
        """Create main comparison sheet."""
        ws = self.workbook.create_sheet('Price Comparison')

        self._create_header_row(ws)
        self._format_data_rows(ws, products)

        # Set up auto-filter
        max_row = len(products) + 1
        ws.auto_filter.ref = f'A1:AA{max_row}'

    def create_raw_data_sheet(self, raw_products: List[Dict]):
        """Create raw data sheet for reference."""
        ws = self.workbook.create_sheet('Raw Data', 0)

        # Get all unique fields from products
        all_fields = set()
        for product in raw_products:
            all_fields.update(product.keys())

        fields = sorted(list(all_fields))

        # Write headers
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='595959', end_color='595959', fill_type='solid')

        # Write data
        for row_idx, product in enumerate(raw_products, start=2):
            for col_idx, field in enumerate(fields, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = product.get(field)

                # Convert complex types to strings
                if isinstance(value, (dict, list)):
                    value = str(value)[:100]  # Truncate to 100 chars
                elif value is None:
                    value = 'N/A'

                cell.value = value

        # Auto-fit columns
        for col_idx, field in enumerate(fields, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Freeze top row
        ws.freeze_panes = 'A2'

    def save(self):
        """Save workbook to file."""
        # Create output directory if needed
        os.makedirs(os.path.dirname(self.output_path) or '.', exist_ok=True)

        self.workbook.save(self.output_path)
        print(f"\n✓ Excel file saved: {self.output_path}")

        # Print file size
        file_size = os.path.getsize(self.output_path)
        print(f"  File size: {file_size / 1024:.1f} KB")

    def create_gap_analysis_sheet(self, sheet_name: str, base_label: str,
                                 gap_rows: List[Dict], summary: Dict):
        """Create one cross-platform gap analysis sheet (base platform
        vs Noon). Called once per comparison (universe, Jarir, Extra,
        Amazon.sa) to produce separate sheets."""
        ws = self.workbook.create_sheet(sheet_name)

        # Summary block at top
        summary_lines = [
            (f'Total {base_label} products (base):', summary.get('total_base_products')),
            ('Exact match on Noon:', f"{summary.get('exact_match_count')} ({summary.get('exact_match_pct')}%)"),
            ('Similar product available on Noon:', f"{summary.get('similar_available_count')} ({summary.get('similar_available_pct')}%)"),
            ('Not available on Noon at all:', f"{summary.get('not_available_count')} ({summary.get('not_available_pct')}%)"),
        ]
        for i, (label, value) in enumerate(summary_lines, start=1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)

        header_row = len(summary_lines) + 2

        columns = [
            ('master_sku', 'Master SKU'), ('title', 'Title'), ('category', 'Category'),
            ('brand', 'Brand'), ('model_name', 'Model'), ('processor', 'Processor'),
            ('processor_full', 'Processor (Full)'), ('ram', 'RAM'), ('storage', 'Storage'),
            ('graphics_card', 'GPU'), ('ai_classification', 'AI'),
            ('available_on', 'Available On'),
            ('base_price', f'{base_label} Price'), ('base_link', f'{base_label} Link'),
            ('compare_status', 'Noon Status'), ('compare_price', 'Noon Price'),
            ('price_diff_vs_compare', 'Price Diff vs Noon'),
            ('compare_similar_product', 'Similar Noon Product'), ('match_confidence', 'Match Confidence'),
            ('compare_link', 'Noon Link'),
        ]

        for col_idx, (field, header) in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        status_fills = {
            'Exact Match': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'Similar Available': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'Not Available': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        }

        for row_idx, row_data in enumerate(gap_rows, start=header_row + 1):
            status = row_data.get('compare_status')
            for col_idx, (field, _) in enumerate(columns, start=1):
                value = row_data.get(field)
                cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else 'N/A')
                if field == 'compare_status' and status in status_fills:
                    cell.fill = status_fills[status]

        ws.freeze_panes = f'A{header_row + 1}'
        ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(columns))}{header_row + len(gap_rows)}'

    @staticmethod
    def merge_data_and_export(unified_products: List[Dict], raw_products: List[Dict],
                             output_path: str, comparisons: Dict[str, Dict] = None):
        """
        Convenience method to create and populate Excel file.

        Args:
            unified_products: List of merged/unified products
            raw_products: List of all raw scraped products
            output_path: Where to save the Excel file
            comparisons: Optional dict of {key: {'rows': [...], 'summary': {...}}}
                from gap_analyzer.py, e.g. {'universe': ..., 'jarir': ...,
                'extra': ..., 'amazon_sa': ...} - one sheet is created per entry
        """
        exporter = ExcelExporter(output_path)
        exporter.create_raw_data_sheet(raw_products)
        exporter.create_comparison_sheet(unified_products)

        if comparisons:
            sheet_specs = [
                ('universe', 'Noon Gap - Universe', 'Universe'),
                ('jarir', 'Noon Gap - Jarir', 'Jarir'),
                ('extra', 'Noon Gap - Extra', 'Extra'),
                ('amazon_sa', 'Noon Gap - Amazon.sa', 'Amazon.sa'),
            ]
            for key, sheet_name, base_label in sheet_specs:
                comparison = comparisons.get(key)
                if comparison:
                    exporter.create_gap_analysis_sheet(
                        sheet_name, base_label, comparison['rows'], comparison['summary']
                    )

        exporter.save()

        return output_path
